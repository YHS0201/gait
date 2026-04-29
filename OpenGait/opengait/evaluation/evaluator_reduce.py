import os
from time import strftime, localtime
import numpy as np
from utils import get_msg_mgr, mkdir

from .metric import mean_iou, cuda_dist, compute_ACC_mAP, evaluate_rank, evaluate_many, cuda_dist_all, euc_dist, evaluate_rank_all_analysis
from .re_rank import re_ranking
import json
from prettytable import PrettyTable, ALL
import csv
import torch.nn.functional as F
import pickle as pkl


def evaluate_reduce_target(data,
                           dataset,
                           metric='euc',
                           dataset_partition='./datasets/Gait3D/Gait3D.json',
                           save_json='show_result/Gait3D.json',
                           enable_save_csv=True,
                           split_eva=True,
                           tb_name='yh_five',
                           probe_num=0,
                           gallery_num=0,
                           *args,
                           **kwargs):
    msg_mgr = get_msg_mgr()
    msg_mgr.log_info(f"dataset_partition is {dataset_partition}")
    features, labels, types, time_seqs = data['embeddings'], data[
        'labels'], data['types'], data['views']
    features = features[:, :, :1]
    msg_mgr.log_info(f"features.shape:{features.shape}")
    msg_mgr.log_info(f"in {__file__},set path is {dataset_partition}")
    probe_sets = json.load(open(dataset_partition, 'rb'))['PROBE_SET']
    # probe_sets = json.load(open(dataset_partition, 'rb'))['TRAIN_PROBE_SET']
    save_json_root = save_json[:-5]
    os.makedirs(save_json_root, exist_ok=True)

    # eva_type_list = []

    # cloth_eva_dict = {}
    # scene_eva_dict = {}
    # dist_eva_dict = {}
    # inout_eva_dict = {}
    # sex_eva_dict = {}
    # time_eva_dict = {}
    # dataset_eva_dict = {}
    #
    # for i in range(len(types)):
    #     if 'noise' in labels[i] or 'query' in types[i]:
    #         continue
    #     # cloth_type, scene_type, dist_type, inout_type = types[i].split('_', 3)
    #     # print(types[i])
    #     if len(types[i].split('_')) == 7:
    #         cloth_type, scene_type, dist_type, inout_type, sex_type, time_type, dataset_type = types[
    #             i].split('_')
    #     elif len(types[i].split('_')) == 8:
    #         cloth_type, scene_type, scene_type2, dist_type, inout_type, sex_type, time_type, dataset_type = types[
    #             i].split('_')
    #         scene_type = scene_type + '_' + scene_type2
    #     else:
    #         continue
    #     cloth_eva_dict.setdefault(cloth_type, []).append(types[i])
    #     # scene_eva_dict.setdefault(scene_type, []).append(types[i])  # commit...for comp
    #     dist_eva_dict.setdefault(dist_type, []).append(types[i])
    #     inout_eva_dict.setdefault(inout_type, []).append(types[i])
    #     sex_eva_dict.setdefault(sex_type, []).append(types[i])
    #     time_eva_dict.setdefault(time_type, []).append(types[i])
    #     dataset_eva_dict.setdefault(dataset_type, []).append(types[i])
    #
    # eva_type_dict = {
    #     **cloth_eva_dict,
    #     **scene_eva_dict,
    #     **dist_eva_dict,
    #     **inout_eva_dict,
    #     **sex_eva_dict,
    #     **time_eva_dict,
    #     **dataset_eva_dict
    # }

    cloth_eva_dict = {}
    day_night_dict = {}
    dataset_dict = {}

    for i in range(len(types)):
        # if 'noise' in labels[i] or 'query' in types[i]:
        #     continue
        types_sp = types[i].split('_')
        if 'jianzhi2503' in labels[i]:
            cloth = types_sp[0]
            dn = types_sp[-2]
            dataset = 'jianzhi2503'
        elif 'yh40' in labels[i]:
            if '0query' in types[i] or 'clothes1' in types[i]:
                cloth = 'NoChange'
            else:
                cloth = 'UpChange'
            if 'night' in types[i]:
                dn = 'night'
            else:
                dn = 'day'
            dataset = 'yh40'
        elif 'jianzhi50' in labels[i]:
            if '0query' in types[i] or 'clothes1' in types[i]:
                cloth = 'NoChange'
            elif 'clothes2' in types[i]:
                cloth = 'DnChange1'
            elif 'clothes3' in types[i]:
                if 'woman' in labels[i]:
                    cloth = 'DnChange2'
                else:
                    cloth = 'UpDnChange1'
            elif 'clothes4' in types[i]:
                if 'woman' in labels[i]:
                    cloth = 'UpDnChange1'
                else:
                    cloth = 'UpChange'
            else:
                cloth = 'UnknClothes'
            dn = 'day'
            dataset = 'jianzhi50'
        elif 'dh30' in labels[i]:
            if '0query' in types[i] or 'clothes1' in types[i]:
                cloth = 'NoChange'
            else:
                cloth = 'UpChange'
            if 'night' in types[i]:
                dn = 'night'
            else:
                dn = 'day'
            dataset = 'dh30'
        elif 'CCGR' in labels[i]:
            if 'NM1' == types[i]:
                continue
            if 'CL' in types[i]:
                cloth = 'OutChange'
            else:
                cloth = 'NoChange'
            dn = 'day'
            dataset = 'CCGR'
        elif 'DT' in labels[i]:
            if '0query' in types[i] or 'clothes1' in types[i]:
                cloth = 'NoChange'
            else:
                cloth = 'UpChange'
            dn = 'day'
            dataset = 'DT'
        elif '250214_dhdata' in labels[i]:
            cloth = 'UnknClothes'
            dn = 'day'
            dataset = '250214_dhdata'
        #else:
        #    cloth = types_sp[0]
        #    dn = 'day'
        #    dataset = 'jianzhi2507'
        else:
            if 'scene' in types[i]:
                cloth = 'Scene'
            elif 'normal' in types[i]:
                cloth = 'Normal'
            else:
                cloth = 'UnknClothes'
            dn = 'day'
            dataset = 'UnknDataset'
        cloth_eva_dict.setdefault(cloth, []).append(types[i])
        day_night_dict.setdefault(dn, []).append(types[i])
        dataset_dict.setdefault(dataset, []).append(types[i])
    eva_type_dict = {
        **cloth_eva_dict,
        **day_night_dict,
        **dataset_dict,
    }

    eva_type_list = list(eva_type_dict.keys())

    result_dict = {}

    if split_eva:
        msg_mgr.log_info(
            f"get probe and gallery num :{probe_num}-{gallery_num}")
        for eva_type in eva_type_list:
            if eva_type == '0query':
                continue
            save_json = os.path.join(save_json_root, f"{eva_type}.json")
            # print(f'eva type is {eva_type}')
            # if eva_type != 'NoChange':
            #     continue
            result = evaluate_type_reduce_enh(eva_type_dict[eva_type],
                                              data,
                                              probe_sets,
                                              msg_mgr,
                                              metric,
                                              save_json,
                                              probe_num=probe_num,
                                              gallery_num=gallery_num)
            result_dict[eva_type] = result
    all_save_json = os.path.join(save_json_root, "all.json")
    result = evaluate_type_reduce_enh('all',
                                      data,
                                      probe_sets,
                                      msg_mgr,
                                      metric,
                                      all_save_json,
                                      probe_num=probe_num,
                                      gallery_num=gallery_num)
    result_dict['all'] = result

    table_list = ['type'] + list(result_dict['all'].keys())

    table = PrettyTable(table_list)
    table.hrules = ALL
    for k, v in result_dict.items():
        data_list = [k]
        for v_k, v_v in v.items():
            data_list.append(round(v_v, 3))
        # print(table_list)
        # print(data_list)
        table.add_row(data_list)
    msg_mgr.log_info(f'\n {table}')

    if enable_save_csv:
        # 保存成csv，方便记录
        csv_columns = ['all'] + eva_type_list
        index_columns = result_dict['all'].keys()
        csv_data = {}
        for data_type in csv_columns:
            if data_type == '0query':
                continue
            for index in index_columns:
                csv_data[f"{data_type}_{index}"] = result_dict[data_type][
                    index]

        filename = 'csv/yh_jianzhi_record_test.csv'

        os.makedirs(os.path.dirname(filename), exist_ok=True)
        with open(filename, 'a', newline='') as file:
            writer = csv.DictWriter(file, fieldnames=csv_data.keys())
            if file.tell() == 0:
                writer.writeheader()
            writer.writerow(csv_data)

        print(f"新数据已追加到 {filename}")

    return_result = {}
    for key, value in result_dict['all'].items():
        if key == 'gallery_num':
            continue
        return_result[f'{tb_name}/{key}'] = value

    return return_result


# def evaluate_type_reduce(eva_type,
#                          data,
#                          probe_sets,
#                          msg_mgr,
#                          metric,
#                          save_json,
#                          probe_num=0,
#                          gallery_num=0):
#
#     features, labels, cams, time_seqs = data['embeddings'], data[
#         'labels'], data['types'], data['views']
#     # features = features[:, :, ::4]
#     probe_mask = []
#     gallery_mask = []
#     all_gallery_mask = []
#     # print(probe_sets)
#     probe_id_num_record = {}
#     gallery_id_num_record = {}
#     for id, ty, sq in zip(labels, cams, time_seqs):
#         probe_id_num_record.setdefault(id, 0)
#         if '-'.join([id, ty, sq]) in probe_sets:
#             if probe_num > 0 and probe_id_num_record[id] < probe_num:
#                 probe_id_num_record[id] += 1
#                 probe_mask.append(True)
#             elif probe_num == 0:
#                 probe_id_num_record[id] += 1
#                 probe_mask.append(True)
#             else:
#                 probe_mask.append(False)
#         else:
#             probe_mask.append(False)
#         # if eva_type == 'all':
#         #     continue
#         if isinstance(eva_type, list):
#             if ty in eva_type:
#                 gallery_id_num_record.setdefault(id, 0)
#                 if gallery_num > 0 and gallery_id_num_record[id] < gallery_num:
#                     gallery_id_num_record[id] += 1
#                     gallery_mask.append(True)
#                 elif gallery_num == 0:
#                     gallery_id_num_record[id] += 1
#                     gallery_mask.append(True)
#                 else:
#                     gallery_mask.append(False)
#             elif ('noise' in id):
#                 gallery_mask.append(True)
#             elif eva_type == 'all':
#                 gallery_id_num_record.setdefault(id, 0)
#                 if gallery_num > 0 and gallery_id_num_record[
#                         id] < gallery_num and not '-'.join([id, ty, sq
#                                                             ]) in probe_sets:
#                     gallery_id_num_record[id] += 1
#                     gallery_mask.append(True)
#                 elif gallery_num == 0 and not '-'.join([id, ty, sq
#                                                         ]) in probe_sets:
#                     gallery_id_num_record[id] += 1
#                     gallery_mask.append(True)
#                 else:
#                     gallery_mask.append(False)
#             else:
#                 gallery_mask.append(False)
#         else:
#             if ty == eva_type:
#                 gallery_id_num_record.setdefault(id, 0)
#                 if gallery_num > 0 and gallery_id_num_record[id] < gallery_num:
#                     gallery_id_num_record[id] += 1
#                     gallery_mask.append(True)
#                 elif gallery_num == 0:
#                     gallery_id_num_record[id] += 1
#                     gallery_mask.append(True)
#                 else:
#                     gallery_mask.append(False)
#             elif 'noise' in id:
#                 gallery_mask.append(True)
#             elif eva_type == 'all':
#                 gallery_id_num_record.setdefault(id, 0)
#                 if gallery_num > 0 and gallery_id_num_record[
#                         id] < gallery_num and not '-'.join([id, ty, sq
#                                                             ]) in probe_sets:
#                     gallery_id_num_record[id] += 1
#                     gallery_mask.append(True)
#                 elif gallery_num == 0 and not '-'.join([id, ty, sq
#                                                         ]) in probe_sets:
#                     gallery_id_num_record[id] += 1
#                     gallery_mask.append(True)
#                 else:
#                     gallery_mask.append(False)
#             else:
#                 gallery_mask.append(False)
#
#     # if eva_type == 'all':
#     #     gallery_mask = ~np.array(probe_mask)
#     probe_features = features[probe_mask]
#     gallery_features = features[gallery_mask]
#
#     probe_lbls = np.asarray(labels)[probe_mask]
#     gallery_lbls = np.asarray(labels)[gallery_mask]
#
#     results = {}
#     msg_mgr.log_info(f"The test metric you choose is {metric}.")
#     dist = cuda_dist(probe_features, gallery_features, metric).cpu().numpy()
#     # dist = euc_dist(probe_features, gallery_features)
#     # 拍平并保存
#     np.save("./dist.npy", dist)
#
#     # save_dict = {
#     #     'labels': data['labels'],
#     #     'types': data['types'],
#     #     'views': data['views'],
#     #     'features': data['embeddings']
#     # }
#     # with open('debug_data.pkl', 'wb') as f:
#     #     pkl.dump(save_dict, f)
#
#
#     try:
#         cmc, all_AP, all_INP = evaluate_rank(dist, probe_lbls, gallery_lbls)
#
#         mAP = np.mean(all_AP)
#         mINP = np.mean(all_INP)
#         max_len = min(50, len(cmc))
#         for i, r in enumerate([1, 5, 10, max_len]):
#             results['Rank-{}'.format(50 if i == 3 else r)] = cmc[r - 1] * 100
#         results['mAP'] = mAP * 100
#         results['mINP'] = mINP * 100
#         results['pids'] = len(gallery_id_num_record)
#         results['query'] = sum([v if k in gallery_id_num_record else 0 for k, v in probe_id_num_record.items()])
#         results['target'] = sum([v if k in gallery_id_num_record else 0 for k, v in gallery_id_num_record.items()])
#         print(results['pids'], results['query'], results['target'])
#         results['gallery_num'] = gallery_features.shape[0]
#     except:
#         for r in [1, 5, 10, 50]:
#             results['Rank-{}'.format(r)] = 0
#         results['mAP'] = 0
#         results['mINP'] = 0
#         results['query'] = 0
#         results['pids'] = 0
#         results['target'] = 0
#         results['gallery_num'] = 0
#
#     # print_csv_format(dataset_name, results)
#     # msg_mgr.log_info(results)
#
#     # write_results(dist,
#     #               probe_mask,
#     #               labels,
#     #               cams,
#     #               time_seqs,
#     #               save_json,
#     #               gallery_mask=gallery_mask)
#
#     return results


def evaluate_type_reduce(eva_type,
                         data,
                         probe_sets,
                         msg_mgr,
                         metric,
                         save_json,
                         probe_num=0,
                         gallery_num=0):
    features, labels, cams, time_seqs = data['embeddings'], data[
        'labels'], data['types'], data['views']
    # features = features[:, :, ::4]
    probe_mask = []
    gallery_mask = []
    all_gallery_mask = []
    # print(probe_sets)
    probe_id_num_record = {}
    gallery_id_num_record = {}

    for id, ty, sq in zip(labels, cams, time_seqs):
        probe_id_num_record.setdefault(id, 0)
        if '-'.join([id, ty, sq]) in probe_sets:
            if probe_num > 0 and probe_id_num_record[id] < probe_num:
                probe_id_num_record[id] += 1
                probe_mask.append(True)
            elif probe_num == 0:
                probe_id_num_record[id] += 1
                probe_mask.append(True)
            else:
                probe_mask.append(False)
            gallery_mask.append(False)
        else:
            probe_mask.append(False)
            gallery_mask.append(True)
 
    probe_features = features[probe_mask]
    gallery_features = features[gallery_mask]

    probe_lbls = np.asarray(labels)[probe_mask]
    gallery_lbls = np.asarray(labels)[gallery_mask]
    gallery_cams = np.asarray(cams)[gallery_mask]
    gallery_time_seqs = np.asarray(time_seqs)[gallery_mask]
    
    
    gallery_pid_mask = np.zeros((len(probe_lbls), len(gallery_lbls)))
    
    gallery_number = 0
    gallery_pid_number = {}
    target_number = 0
    # np.save('gallery_lbls_ori.npy', gallery_lbls.copy())
    for id, ty, sq in zip(gallery_lbls, gallery_cams, gallery_time_seqs):
        # 找出 id 在 probe_lbls 中的索引
        idx = np.where(probe_lbls == id)[0]
        if len(idx) == 0:
            gallery_number += 1
            continue
        if not isinstance(eva_type, list):
            eva_type = [eva_type]
        if eva_type[0] != 'all' and ty not in eva_type:
            gallery_pid_mask[idx, gallery_number] = 1
            gallery_lbls[gallery_number] = 'noise'
        else:
            if id not in gallery_pid_number:
                gallery_pid_number[id] = 0
            if gallery_num != 0 and gallery_pid_number[id] >= gallery_num:
                gallery_pid_mask[idx, gallery_number] = 1
                gallery_lbls[gallery_number] = 'noise'
            else:
                target_number += 1
                gallery_pid_number[id] += 1
        gallery_number += 1
    
    # 因为上一步将 gallery 进行了过滤，有可能出现某些 query 在 gallery 中没有对应的 id, 这部分数据会导致指标异常
    probe_new_mask = np.ones_like(probe_lbls)
    for i, probe_label in enumerate(probe_lbls):
        if probe_label not in gallery_lbls:
            probe_new_mask[i] = 0
    probe_new_mask = probe_new_mask.astype(np.int8)
    probe_features = probe_features[probe_new_mask == 1]
    probe_lbls = probe_lbls[probe_new_mask == 1]
    gallery_pid_mask = gallery_pid_mask[probe_new_mask == 1]


    # gallery_pid_mask = gallery_pid_mask.squeeze(0)
    # np.save('probe_lbls.npy', probe_lbls)
    # np.save('gallery_lbls.npy', gallery_lbls)
    # np.save('gallery_cams.npy', gallery_cams)
    # np.save('gallery_time_seqs.npy', gallery_time_seqs)

    results = {}
    msg_mgr.log_info(f"The test metric you choose is {metric}.")
    dist = cuda_dist(probe_features, gallery_features, metric).cpu().numpy()
    # dist = euc_dist(probe_features, gallery_features)
    # 拍平并保存
    np.save("./dist.npy", dist)
    
    # 将符合条件的dist设为-1
    dist[gallery_pid_mask == 1] = 2

    # 将对应的gallery_labls 设为 noise
    
    # save_dict = {
    #     'labels': data['labels'],
    #     'types': data['types'],
    #     'views': data['views'],
    #     'features': data['embeddings']
    # }
    # with open('debug_data.pkl', 'wb') as f:
    #     pkl.dump(save_dict, f)

    # try:

    # print(f'type is {eva_type}')

    cmc, all_AP, all_INP = evaluate_rank(dist, probe_lbls, gallery_lbls)
    if cmc is None:
        # print(f'eva_type is {eva_type}')
        print(f'cmc is None')
        for r in [1, 5, 10, 50]:
            results['Rank-{}'.format(r)] = 0
        results['mAP'] = 0
        results['mINP'] = 0
        results['query_pids'] = 0
        results['query_seq_number'] = 0
        results['mean_target_number'] = 0
        results['gallery_num'] = 0
        return results


    mAP = np.mean(all_AP)
    mINP = np.mean(all_INP)
    max_len = min(50, len(cmc))
    for i, r in enumerate([1, 5, 10, max_len]):
        results['Rank-{}'.format(50 if i == 3 else r)] = cmc[r - 1] * 100
    results['mAP'] = mAP * 100
    results['mINP'] = mINP * 100
    results['query_pids'] = len(list(set(probe_lbls)))
    results['query_seq_number'] = dist.shape[0]
    target_pid_num = len(gallery_pid_number.keys())
    results['mean_target_number'] = target_number / target_pid_num if target_pid_num > 0 else 0
    # print(results['query_pids'], results['query_seq_number'], results['mean_target_number'])
    results['gallery_num'] = dist.shape[1]
    print(results)
    # except:
    #     for r in [1, 5, 10, 50]:
    #         results['Rank-{}'.format(r)] = 0
    #     results['mAP'] = 0
    #     results['mINP'] = 0
    #     results['query_pids'] = 0
    #     results['query_seq_number'] = 0
    #     results['mean_target_number'] = 0
    #     results['gallery_num'] = 0

    return results


def evaluate_type_reduce_enh(eva_type,
                             data,
                             probe_sets,
                             msg_mgr,
                             metric,
                             save_json,
                             probe_num=0,
                             gallery_num=0,
                             save_pkl=None,
                             save_excel=None):
    """Enhanced evaluator with per-query ranked gallery trace saved to pkl/excel."""
    features, labels, cams, time_seqs = data['embeddings'], data[
        'labels'], data['types'], data['views']
    probe_mask = []
    gallery_mask = []
    probe_id_num_record = {}

    for id, ty, sq in zip(labels, cams, time_seqs):
        probe_id_num_record.setdefault(id, 0)
        if '-'.join([id, ty, sq]) in probe_sets:
            if probe_num > 0 and probe_id_num_record[id] < probe_num:
                probe_id_num_record[id] += 1
                probe_mask.append(True)
            elif probe_num == 0:
                probe_id_num_record[id] += 1
                probe_mask.append(True)
            else:
                probe_mask.append(False)
            gallery_mask.append(False)
        else:
            probe_mask.append(False)
            gallery_mask.append(True)

    probe_features = features[probe_mask]
    gallery_features = features[gallery_mask]

    probe_lbls = np.asarray(labels)[probe_mask]
    probe_cams = np.asarray(cams)[probe_mask]
    probe_time_seqs = np.asarray(time_seqs)[probe_mask]

    gallery_lbls = np.asarray(labels)[gallery_mask]
    gallery_cams = np.asarray(cams)[gallery_mask]
    gallery_time_seqs = np.asarray(time_seqs)[gallery_mask]

    gallery_pid_mask = np.zeros((len(probe_lbls), len(gallery_lbls)))

    gallery_number = 0
    gallery_pid_number = {}
    target_number = 0
    for id, ty, sq in zip(gallery_lbls, gallery_cams, gallery_time_seqs):
        idx = np.where(probe_lbls == id)[0]
        if len(idx) == 0:
            gallery_number += 1
            continue
        if not isinstance(eva_type, list):
            eva_type = [eva_type]
        if eva_type[0] != 'all' and ty not in eva_type:
            gallery_pid_mask[idx, gallery_number] = 1
            gallery_lbls[gallery_number] = 'noise'
        else:
            if id not in gallery_pid_number:
                gallery_pid_number[id] = 0
            if gallery_num != 0 and gallery_pid_number[id] >= gallery_num:
                gallery_pid_mask[idx, gallery_number] = 1
                gallery_lbls[gallery_number] = 'noise'
            else:
                target_number += 1
                gallery_pid_number[id] += 1
        gallery_number += 1

    probe_new_mask = np.ones_like(probe_lbls)
    for i, probe_label in enumerate(probe_lbls):
        if probe_label not in gallery_lbls:
            probe_new_mask[i] = 0
    probe_new_mask = probe_new_mask.astype(np.int8)
    probe_features = probe_features[probe_new_mask == 1]
    probe_lbls = probe_lbls[probe_new_mask == 1]
    probe_cams = probe_cams[probe_new_mask == 1]
    probe_time_seqs = probe_time_seqs[probe_new_mask == 1]
    gallery_pid_mask = gallery_pid_mask[probe_new_mask == 1]

    results = {}
    msg_mgr.log_info(f"The test metric you choose is {metric}.")
    dist = cuda_dist(probe_features, gallery_features, metric).cpu().numpy()
    np.save("./dist.npy", dist)

    dist[gallery_pid_mask == 1] = 2

    cmc, all_AP, all_INP = evaluate_rank(dist, probe_lbls, gallery_lbls)
    if cmc is None:
        print(f'cmc is None')
        for r in [1, 5, 10, 50]:
            results['Rank-{}'.format(r)] = 0
        results['mAP'] = 0
        results['mINP'] = 0
        results['query_pids'] = 0
        results['query_seq_number'] = 0
        results['mean_target_number'] = 0
        results['gallery_num'] = 0
    else:
        mAP = np.mean(all_AP)
        mINP = np.mean(all_INP)
        max_len = min(50, len(cmc))
        for i, r in enumerate([1, 5, 10, max_len]):
            results['Rank-{}'.format(50 if i == 3 else r)] = cmc[r - 1] * 100
        results['mAP'] = mAP * 100
        results['mINP'] = mINP * 100
        results['query_pids'] = len(list(set(probe_lbls)))
        results['query_seq_number'] = dist.shape[0]
        target_pid_num = len(gallery_pid_number.keys())
        results['mean_target_number'] = target_number / target_pid_num if target_pid_num > 0 else 0
        results['gallery_num'] = dist.shape[1]
        print(results)

    # Save per-query gallery ranking trace for analysis.
    rank_trace = []
    if dist.size > 0:
        sorted_idx = np.argsort(dist, axis=1)
        for qi in range(dist.shape[0]):
            order = sorted_idx[qi].tolist()
            ranked_gallery = []
            for rank, gi in enumerate(order, 1):
                ranked_gallery.append({
                    'rank': rank,
                    'id': str(gallery_lbls[gi]),
                    'ty': str(gallery_cams[gi]),
                    'sq': str(gallery_time_seqs[gi]),
                    'dist': float(dist[qi, gi]),
                    'masked': bool(gallery_pid_mask[qi, gi] == 1)
                })
            rank_trace.append({
                'query': {
                    'id': str(probe_lbls[qi]),
                    'ty': str(probe_cams[qi]),
                    'sq': str(probe_time_seqs[qi])
                },
                'gallery_ranked': ranked_gallery
            })

    if save_pkl is None:
        save_pkl = os.path.splitext(save_json)[0] + "_gallery_rank.pkl"
    save_pkl_dir = os.path.dirname(save_pkl)
    if save_pkl_dir:
        os.makedirs(save_pkl_dir, exist_ok=True)

    save_payload = {
        'eva_type': eva_type,
        'metric': metric,
        'probe_num': probe_num,
        'gallery_num': gallery_num,
        'results': results,
        'rank_trace': rank_trace
    }
    with open(save_pkl, 'wb') as f:
        pkl.dump(save_payload, f)
    msg_mgr.log_info(f"Saved enhanced rank trace pkl to: {save_pkl}")

    if save_excel is None:
        save_excel = os.path.splitext(save_json)[0] + "_gallery_rank.xlsx"
    save_excel_dir = os.path.dirname(save_excel)
    if save_excel_dir:
        os.makedirs(save_excel_dir, exist_ok=True)

    try:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "rank_trace"

        for col_idx, item in enumerate(rank_trace, 1):
            query = item['query']
            ws.cell(
                row=1,
                column=col_idx,
                value='-'.join([query['id'], query['ty'], query['sq']]))

            ranked_gallery = item['gallery_ranked']
            for row_off, g in enumerate(ranked_gallery, 2):
                ws.cell(
                    row=row_off,
                    column=col_idx,
                    value='-'.join([
                        str(g['rank']),
                        str(g['id']),
                        str(g['ty']),
                        str(g['sq']),
                        str(g['dist']),
                        str(g['masked'])
                    ]))
        wb.save(save_excel)
        msg_mgr.log_info(f"Saved enhanced rank trace excel to: {save_excel}")
    except Exception as e:
        msg_mgr.log_info(f"Save excel failed: {e}")

    return results
